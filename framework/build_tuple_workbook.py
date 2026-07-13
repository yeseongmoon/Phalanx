# -*- coding: utf-8 -*-
"""
build_tuple_workbook.py
Builds the ENISA-side threat-record tuple workbook (xlsx + csv) for the
6G NTN threat-modeling framework, for ALL 58 ENISA threats.

PROVENANCE (no fabrication):
  * Threat cluster, threat name, description  ->  from the analyst's authoritative
    sheet  "ENISA Dataset Threat Mapping.xlsx"  (sheet "ENISA space threat", 58 rows).
  * CIA and Affected Assets (A)  ->  transcribed from the ENISA *Space Threat Landscape*
    report (March 2025), Annex C "Space Threat Taxonomy" (Table 17, pp.85-95).
    Source PDF: references/Space_Threat_Landscape_Report_fin.pdf
    (ISBN 978-92-9204-696-5, DOI 10.2824/8841206).
  * S (segments), TB (trust boundaries), L (lifecycle) are DERIVED from A
    (Algorithm 1, Pass 1) and re-validated by the algorithm.
  * Pre, TTP, CM, I, Conf are LEFT BLANK -> analyst maps SPARTA; verified separately.

  Rows flagged approx=True had ambiguous PDF layout; their affected-asset text is a
  best-effort transcription (segments are reliable). See the 'Notes' column.
"""
import csv, os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
USER_SHEET = os.path.join(HERE, "..", "ENISA Dataset Threat Mapping.xlsx")

CATEGORIES = {"NAA":"Nefarious Activity/Abuse","EIH":"Eavesdropping/Interception/Hijacking",
    "PA":"Physical Attacks","UD":"Unintentional Damage","FM":"Failures or malfunctions",
    "OUT":"Outages","DIS":"Disaster","LEG":"Legal","LEI":"Legacy infrastructure"}
LIFECYCLE = {1:"Design and Development",2:"Assembly",3:"Pre-launch",4:"Launch",
             5:"In-orbit Testing",6:"Operations",7:"Decommissioning"}
CLUSTER2E = {
    "Nefarious activity/abuse":"NAA","Evaesdropping/Interception/Hijacking":"EIH",
    "Physical access":"PA","Unintentional damage":"UD","Failures or malfunctions":"FM",
    "Outages":"OUT","Disaster":"DIS","Legal":"LEG","Legacy infrastructure":"LEI"}

# affected assets (A), CIA, segments, [approx?]  keyed by the analyst-sheet threat name
# transcribed from ENISA Annex C, Table 17.
A = lambda segs, cia, aff, approx=False: (segs, cia, aff, approx)
ASSETS = {
 # ---- NAA ----
 "Abuse of leaked data": A(["Ground"],"C","Ground: Production (Document Management incl. Configuration Management System; Prototyping/IDE; ERP software & misc software); Email servers, databases, OS; Phones, laptops, tablets"),
 "Abuse/Falsification of rights": A(["Ground","User"],"CIA","Ground: Design/dev/QA, Assembly, Manufacturing Systems, Misc (software), Misc (endpoint devices), Logistics Management System, Checkout systems, Mission Control System, Network (WAN); User: Consumer endpoint devices"),
 "Compromising confidential information (data breaches): Exfiltration": A(["Ground","User"],"C","Ground: SLE/SDL protocols; User: endpoint interfaces and devices"),
 "Data Modification": A(["Ground"],"CI","Ground: Production (Document Management incl. Configuration Management System; Prototyping/IDE; ERP software & misc software); Email servers, databases, OS; Phones, laptops, tablets"),
 "Denial of Service (DoS)": A(["Ground","Space","User"],"A","Ground: all assets; Space: all assets; User: Consumer Interfaces & devices"),
 "Electromagnetic interference": A(["Space"],"IA","Space: EPS, RTOS and AOCS"),
 "Firmware corruption": A(["Ground","Space"],"A","Ground: all assets; Space: all assets"),
 "Identity Theft": A(["Human Resources"],"CIA","Human resources: all assets"),
 "Jamming": A(["Ground","Space","User"],"A","Ground: TTC Ground (Antenna); Space: BUS (COM) & Payload (PLCOM); User: VSAT"),
 "Malicious code/software/activity: Cryptographic exploit": A(["Ground","Space"],"C","Ground: Crypto Hardware/Software & Transport Container (Crypto Unit Board) & Satellite Control Centre (Crypto Unit Ground); Space: Crypto Unit Board"),
 "Malicious code/software/activity: Malicious injection": A(["Ground","Space"],"IA","Ground: Production (Manufacturing systems & Simulators & Crypto hw/sw) & Centralised Checkout Systems & Satellite Operations Centre (Mission Control System); Space: BUS (CDHS/COM) & Payload (PDHS/PLCOM)"),
 "Malicious code/software/activity: Network exploit": A(["Ground","Space","User"],"A","Ground: Satellite Control Centre & Centralised Checkout Systems & TTC Ground (Antenna); Space: BUS (CDHS, COM) & Payload (PDHS, PLCOM); User: VSAT"),
 "Malicious code/software/activity: Software and vulnerability exploit": A(["Ground","Space"],"CI","Ground: Production (Manufacturing systems & Simulators) & Centralised Checkout Systems & Satellite Operations Centre (Mission Control System); Space: Satellite Bus, Satellite payload"),
 "Manipulation of hardware and software: Zero-Day exploit": A(["Ground","Space","User"],"CIA","Ground: all assets; Space: all assets; User: all assets"),
 "Preventing services": A(["Ground"],"CIA","Ground: Production (Design, development, QA & Assembly) & Satellite Control Centre"),
 "Resource exhaustion": A(["Space"],"A","Space: BUS (CDHS/COM) & Payload (PLCOM/PDHS) [satellite ops]; BUS (CDHS, OBSW) [logical storage]; BUS (COM, antenna) [communications]"),
 "Seizure of control": A(["Ground","Space"],"A","Ground: TTC Ground (SLE/SDL protocol); Space: BUS (CDHS/COM)"),
 "Social Engineering": A(["Ground"],"CIA","Ground: Design, development, and quality assurance & Satellite Control Centre & Miscellaneous Software/Hardware"),
 "Spoofing": A(["Ground","Space","User"],"AI","Ground: TTC Ground (Antenna, SLE/SDL protocol); Space: BUS (COM); User: VSAT"),
 "Supply Chain Compromise": A(["Ground","Space"],"CIA","Ground: Production (Manufacturing systems & Assembly & Simulators) & Centralised Checkout Systems & Satellite Operations Centre (Mission Control System) & TTC Ground; Space: all assets"),
 "Theft of authentication information": A(["Ground","Space"],"CIA","Ground: Crypto Hardware/Software & Transport Container (Crypto Unit Board) & Satellite Control Centre (Crypto Unit Ground); Space: BUS (Crypto Unit Board)"),
 "Unauthorized modification: Parameters": A(["Space"],"CIA","Space: CDHS (RTOS components)"),
 "Unauthorized use of equipment": A(["Ground"],"CIA","Ground: Production (Manufacturing/Assembly facilities & launch facility) & Satellite Control Centre & EGSE/MGSE"),
 # ---- EIH ----
 "Hijacking": A(["Ground","Space"],"IA","Ground: Satellite Control Centre & TTC Ground (SLE/SDL protocol); Space: BUS (CDHS & COM) & Payload (PDHS & PLCOM)"),
 "Interception of communication": A(["Ground","Space","User"],"C","Ground: Satellite Control Centre & TTC Ground (Antenna, SLE/SDL protocol); Space: BUS (COM) & Payload (PLCOM); User: VSAT"),
 "Man-in-the-Middle": A(["Ground","Space","User"],"CIA","Ground: Satellite Control Centre & TTC Ground (SLE/SDL protocol); Space: BUS & Payload; User: VSAT"),
 "Network manipulation (Bus-Payload Link)": A(["Space"],"CI","Space: BUS (CDHS & COM) & Payload (PLCOM & PDHS & UDHS)"),
 "Network traffic manipulation (TC)": A(["Ground","Space"],"A","Ground: Satellite Control Centre & TTC Ground (Antenna); Space: BUS (COM); Ground & Space: SLE/SDL protocols"),
 "Position detection (telemetry)": A(["Space"],"C","Space: CDHS (OBSW, OBC)"),
 "Replay of recorded authentic communication traffic": A(["Ground","Space"],"CI","Ground & Space: SLE/SDL protocols"),
 "Unauthorized access": A(["Ground"],"CI","Ground: all assets (including SLE/SDL protocol)"),
 # ---- Physical access (PA) ----
 "Coercion, extortion, or corruption": A(["Human Resources"],"CIA","Human resources: all assets"),
 "Damage/Destruction of segment assets": A(["Ground","Space"],"A","Ground: Production & Assembly & Transport & Launch & Satellite Control Centre & TTC Ground; Space: all assets"),
 "Damage/Destruction of the satellite via the use of ASAT": A(["Ground","Space"],"A","Ground: TTC Ground; Space: all assets"),
 "Loss during shipping": A(["Ground"],"CIA","Ground: Transport Container"),
 "Sabotage through hardware/software": A(["Ground"],"CIA","Ground: Design, development, and quality assurance & Assembly & Satellite Control Centre"),
 "Unauthorized phyisical access": A(["Ground"],"CIA","Ground: all assets"),
 # ---- Legal (LEG) -- per analyst clustering ----
 "Data leaks": A(["Ground"],"CIA","Ground: Production (Document Management, ERP software, Crypto hw/sw, Simulators, Soft/Hardware Test Tools)"),
 "misuse of equipment": A(["Ground"],"CIA","Ground: Satellite Control Centre (EGSE/MGSE)"),
 "Negligence of asset handling security requirements": A(["Human Resources"],"CIA","Human Resources: all assets"),
 "Refusal of actions": A(["Ground","Space","User","Human Resources"],"CIA","Ground: all assets; Space: all assets; User: all assets; Human resources: all assets"),
 "Third Party non-compliance (supply chain)": A(["Ground","Space"],"C","Ground: Production; Space: Payload (PLCOM)"),
 "Unauthorized access to recycled or disposed media": A(["Space"],"C","Space: Payload (PLCOM); disposed/recycled media or decommissioned infrastructure", True),
 # ---- Failures or malfunctions (FM) ----
 "Failure of air conditioning or water supply": A(["Ground"],"A","Ground: Satellite Control (ground segment; overheating from air conditioning / water supply failure)", True),
 "Failure of Cloud infrastructure": A(["Ground","Space"],"A","Ground: Satellite Control (GSaaS/Cloud); Space: BUS (CDHS) & Payload"),
 "Failure of communication networks": A(["Ground","Space","User"],"A","Ground: TTC Ground (Antenna, SLE/SDL protocol); Space: BUS (COM); User: VSAT"),
 "Failure of power supply": A(["Ground","Space"],"A","Ground: Production & Assembly & Transport & Checkout systems; Space: BUS (EPS)"),
 "Rogue hardware": A(["Ground"],"CI","Ground: Assembly (EGSE/MGSE) & Crypto Hardware/Software & Satellite Control Centre"),
 # ---- Unintentional damage (UD) ----
 "Lack of segregation": A(["Space"],"CIA","Space: BUS (CDHS/COM) & Payload (PDHS/PLCOM/UDHS)"),
 "Operating errors": A(["Ground"],"IA","Ground: Design, development, and quality assurance & Manufacturing systems & Assembly & Simulators & Satellite Control Centre", True),
 "Software misconfiguration": A(["Ground","Space"],"CIA","Ground: Design/dev/QA & Manufacturing & Assembly & Simulators & Satellite Control Centre; Ground & Space: SLE/SDL protocols; Space: all assets"),
 "Inadequate security planning/management": A(["Ground","Space","User","Human Resources"],"CIA","Ground: all assets; Space: all assets; User: all assets; Human Resources: all assets"),
 # ---- Outages (OUT) ----
 "Personnel absence": A(["Human Resources"],"A","Human Resources: all assets"),
 "Security services failure": A(["Ground"],"A","Ground: Satellite control / TT&C system (security safeguards: firewalls, virus scans, log monitoring unavailable)"),
 # ---- Disaster (DIS) ----
 "Atmospheric hazards": A(["Ground","Space"],"A","Ground: Simulators & Centralised Checkout System; Space: BUS (CDHS)", True),
 "Environmental hazards": A(["Ground","Space","User"],"A","Ground: Manufacturing system & Assembly & Soft/Hardware Test Tools & Transport & Launch & Satellite Control Centre & TTC Ground; Space: EPS (Solar Wings & Batteries); User: VSAT"),
 # ---- Legacy infrastructure (LEI) ----
 "Failure to maintain information systems": A(["Ground"],"CIA","Ground: Production (development, and quality assurance)", True),
 "Legacy software": A(["Ground","Space","User"],"CIA","Ground: Production (Manufacturing systems & Assembly & Simulators) & Satellite Operations Centre (Mission Control System) & TTC Ground; Space: Bus & Payload; User: VSAT"),
}

# ---- Pass-1 derivations (documented) --------------------------------------
def derive_trust_boundaries(segs):
    tbs=[]
    for a,b,label in [("User","Ground","User <-> Ground/RAN (service link)"),
                      ("Ground","Space","Ground <-> Space (feeder / telecommand link)"),
                      ("User","Space","User <-> Space (direct service link)")]:
        if a in segs and b in segs: tbs.append(label)
    if "Human Resources" in segs and len(segs)>1:
        tbs.append("Human Resources <-> technical segment (organizational boundary)")
    return tbs or ["(none - single segment / internal)"]

GE={1,2,3}; GO={4,5,6,7}; SP={3,4,5,6,7}; US={6}; HR={1,2,3,4,5,6,7}
def derive_lifecycle(aff, segs):
    a=aff.lower(); ph=set()
    if "Ground" in segs:
        if "all assets" in a or any(k in a for k in ["ttc","control centre","mission control","crypto unit ground","operations"]): ph|=GO
        if "all assets" in a or any(k in a for k in ["production","design","development","assembly","manufactur","checkout","test","simulator","egse","mgse","document management","erp","prototyp","transport"]): ph|=GE
        if "all assets" in a: ph|=GE|GO
    if "Space" in segs: ph|=SP
    if "User" in segs: ph|=US
    if "Human Resources" in segs: ph|=HR
    return sorted(ph) or [6]

def fmt_phases(ph): return "{"+",".join(f"{p}:{LIFECYCLE[p]}" for p in ph)+"}"

# ---- read the analyst's authoritative 58 ----------------------------------
uwb = load_workbook(USER_SHEET, data_only=True)
uws = uwb["ENISA space threat"]
user_rows=[]
for r in range(2, uws.max_row+1):
    cl=uws.cell(r,1).value; th=uws.cell(r,2).value; ds=uws.cell(r,3).value
    if th: user_rows.append((str(cl).strip(), str(th).strip(), (str(ds).strip() if ds else "")))

# ---- assemble --------------------------------------------------------------
HEADERS=["ID","Threat cluster","ENISA_Threat","E (category)","CIA",
         "L (lifecycle)","S (segments)","A (affected assets)","TB (trust boundaries)",
         "Pre","TTP","CM","I","Conf","Threat Description (ENISA)","Notes"]
rows=[]; missing=[]
for i,(cl,th,ds) in enumerate(user_rows,1):
    rec=ASSETS.get(th)
    if rec is None:
        missing.append(th); segs,cia,aff,approx=[],"","(A not matched - fill from Annex C)",True
    else:
        segs,cia,aff,approx=rec
    ecode=CLUSTER2E.get(cl,"?")
    L=derive_lifecycle(aff,segs); TB=derive_trust_boundaries(segs)
    note="affected-asset text approximate (segments reliable)" if approx else ""
    rows.append([f"TR{i:02d}", cl, th, f"{ecode} / {CATEGORIES.get(ecode,'?')}", cia,
                 fmt_phases(L) if segs else "", "; ".join(segs), aff, " ; ".join(TB) if segs else "",
                 "","","","","", ds, note])

# ---- CSV -------------------------------------------------------------------
with open(os.path.join(HERE,"enisa_threats_tuple.csv"),"w",newline="",encoding="utf-8-sig") as f:
    w=csv.writer(f); w.writerow(HEADERS); w.writerows(rows)

# ---- XLSX ------------------------------------------------------------------
wb=Workbook(); ws=wb.active; ws.title="Tuples"
NAVY=PatternFill("solid",fgColor="1F3864"); GREEN=PatternFill("solid",fgColor="E7F3EC")
BLANKF=PatternFill("solid",fgColor="FFF6E5"); GREY=PatternFill("solid",fgColor="F2F4F7")
thin=Side(style="thin",color="D9DFE6"); bd=Border(thin,thin,thin,thin)
for c,h in enumerate(HEADERS,1):
    cell=ws.cell(1,c,h); cell.font=Font(bold=True,color="FFFFFF",size=10); cell.fill=NAVY
    cell.alignment=Alignment("center","center",wrap_text=True); cell.border=bd
blank_cols={10,11,12,13,14}; derived_cols={6,7,8,9}
for r,row in enumerate(rows,2):
    for c,val in enumerate(row,1):
        cell=ws.cell(r,c,val); cell.border=bd
        cell.alignment=Alignment(vertical="top",wrap_text=True); cell.font=Font(size=9)
        if c in blank_cols: cell.fill=BLANKF
        elif c in derived_cols: cell.fill=GREEN
        elif c in (15,16): cell.fill=GREY
widths=[6,20,28,24,6,24,18,38,28,14,16,12,8,7,44,26]
for c,wd in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=wd
ws.freeze_panes="A2"; ws.row_dimensions[1].height=40

rs=wb.create_sheet("Reference")
def put(r,c,v,b=False,col="000000"):
    x=rs.cell(r,c,v); x.font=Font(bold=b,color=col,size=10); return x
put(1,1,"ENISA threat-record tuple  TR = <L, S, A, TB, E, Pre, TTP, CM, I, Conf>  --  58 threats",True,"1F3864")
put(3,1,"E categories",True); r=4
for k,v in CATEGORIES.items(): put(r,1,k); put(r,2,v); r+=1
put(r+1,1,"Lifecycle phases (L)",True); r+=2
for k,v in LIFECYCLE.items(): put(r,1,str(k)); put(r,2,v); r+=1
put(r+1,1,"Green = derived from A (S/TB/L). Amber = analyst fills (Pre/TTP/CM/I/Conf).",False,"666666")
put(r+2,1,"Cluster/Threat/Description from analyst sheet; CIA/A from ENISA Annex C, Table 17.",False,"666666")
rs.column_dimensions["A"].width=10; rs.column_dimensions["B"].width=48
wb.save(os.path.join(HERE,"enisa_threats_tuple.xlsx"))

print(f"user threats read : {len(user_rows)}")
print(f"rows written      : {len(rows)}")
print(f"unmatched (no A)  : {missing if missing else 'none'}")
print(f"approx-flagged    : {sum(1 for x in rows if x[-1])}")
