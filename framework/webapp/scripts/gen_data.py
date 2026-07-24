# -*- coding: utf-8 -*-
"""
gen_data.py -- Python owns the threat model. This emits, from the xlsx:
  src/data/threats.json       -- the 58 ENISA threats (id, cluster, name, E, cia, A, desc)
  test/parity-fixture.json    -- the Python engine's output per threat (segs, L, rules),
                                 used by the vitest parity test to keep the TS engine honest.

Run from framework/webapp:  ../pptenv/bin/python scripts/gen_data.py
(or point --xlsx / --engine at the framework dir).
"""
import os, sys, json
HERE = os.path.dirname(os.path.abspath(__file__))
WEBAPP = os.path.dirname(HERE)
FRAMEWORK = os.path.dirname(WEBAPP)
sys.path.insert(0, FRAMEWORK)                       # import ntn_threat_algorithm
import ntn_threat_algorithm as ENG
from openpyxl import load_workbook

XLSX = os.path.join(FRAMEWORK, "enisa_threats_tuple.xlsx")
ws = load_workbook(XLSX, data_only=True)["Tuples"]
threats = []
for r in range(2, ws.max_row + 1):
    if not ws.cell(r,1).value: continue
    threats.append({
        "id": ws.cell(r,1).value, "cluster": ws.cell(r,2).value or "",
        "name": ws.cell(r,3).value or "", "E": (ws.cell(r,4).value or "").split("/")[0].strip(),
        "cia": ws.cell(r,5).value or "", "A": ws.cell(r,8).value or "",
        "desc": ws.cell(r,15).value or "",
    })

# Merge authored Pass-2 (SPARTA) mappings from framework/mappings.json, if present, so the
# tool renders TTP/CM/Pre/I/Conf for mapped threats. Pass-1 fields stay engine-derived; this
# only fills the analyst-authored fields. Structured chains are flattened to display strings.
MAP = os.path.join(FRAMEWORK, "mappings.json")
if os.path.exists(MAP):
    mp = json.load(open(MAP, encoding="utf-8"))
    n_mapped = 0
    for t in threats:
        e = mp.get(t["id"])
        if not isinstance(e, dict) or "ttp" not in e:   # skip _about / _schema and unmapped ids
            continue
        if e.get("pre"):  t["pre"] = e["pre"]
        t["ttp"] = " → ".join(f"{x['id']} {x['name']}" for x in e["ttp"])
        if e.get("cm_sparta"): t["cm"] = " · ".join(f"{c['id']} {c['name']}" for c in e["cm_sparta"])
        if e.get("impact"): t["impact"] = e["impact"]
        c = e.get("conf")
        if isinstance(c, dict):
            if c.get("grade"): t["conf"] = c["grade"]
        elif c:
            t["conf"] = str(c)
        t["map"] = e   # structured record for the rich card render
        n_mapped += 1
    print("mappings.json merged:", n_mapped, "threat(s)")

fixture = {}
for r in ENG.run(ENG.load_rows(XLSX)):
    fixture[r["id"]] = {"segs": r["dS"], "L": r["dL"], "rules": r["rules"],
                        "nature": r["nature"], "arm": ("N/A" if r["arm"].startswith("N/A") else "SPARTA"),
                        "prim": r["prim"], "tb": r["dTB"]}

os.makedirs(os.path.join(WEBAPP, "src", "data"), exist_ok=True)
os.makedirs(os.path.join(WEBAPP, "test"), exist_ok=True)
json.dump(threats, open(os.path.join(WEBAPP,"src","data","threats.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(fixture, open(os.path.join(WEBAPP,"test","parity-fixture.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
print("threats.json:", len(threats), "| parity-fixture.json:", len(fixture))
