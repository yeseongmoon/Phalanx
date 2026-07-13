#!/usr/bin/env python3
"""Generate a compact ENISA control-framework dataset for the webapp.

Source (authoritative): ENISA — Space Threat Landscape, Control Framework.
  Repo:  https://github.com/enisaeu/Space-Threat-Landscape
  File:  Control Framework.xlsx  (sheet 'Control Framework_Threat based')

ENISA maps each threat -> a set of security controls, every control cross-referenced
to established frameworks (ISO 27001, NIST CSF/IRs, NIS2, NASA BPG, SPARTA, BSI, METI).
We compile: the control catalogue + threat->controls + cluster->controls indexes.

Run:  ../../pptenv/bin/python scripts/gen_enisa_controls.py
Writes: src/data/enisa_controls.json
"""
import openpyxl, json, os, collections

HERE = os.path.dirname(__file__)
XLSX = os.path.join(HERE, "vendor", "ENISA_Control_Framework.xlsx")
OUT = os.path.join(HERE, "..", "src", "data", "enisa_controls.json")
SHEET = "Control Framework_Threat based"


# ENISA cluster names (incl. their own typos) -> canonical ENISA category code E
CLUSTER_TO_E = {
    "nefarious activity/abuse": "NAA",
    "eavesdropping/interception/hijacking": "EIH",
    "physical attacks": "PA",
    "unintentional damage": "UD", "unintention damage": "UD",
    "failures or malfunctions": "FM",
    "outages": "OUT",
    "disaster": "DIS",
    "legal": "LEG",
    "legacy infrastructure": "LEI",
}


# normalise ENISA cluster names (they carry typos / are sometimes combined with '/')
TYPO = {"access managemen": "access management", "accessc management": "access management",
        "rsik management": "risk management", "managemen": "management"}


def cluster_parts(s):
    out = []
    for p in str(s or "").lower().split("/"):
        p = p.strip()
        p = TYPO.get(p, p)
        if p:
            out.append(p)
    return out


def split(v):
    return [x.strip() for x in str(v or "").split("\n") if x.strip()]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ENISA's own SPARTA-analysis sheet maps SPARTA countermeasure themes -> ENISA control clusters.
    # (theme == SPARTA countermeasure name; 88/90 match the STIX dataset exactly.)
    cluster2themes = collections.defaultdict(set)
    for r in wb["SPARTA"].iter_rows(min_row=8, values_only=True):
        cl, theme = (r[2] if len(r) > 2 else None), (r[5] if len(r) > 5 else None)
        theme = str(theme or "").strip().lower()
        if not theme:
            continue
        for cp in cluster_parts(cl):
            cluster2themes[cp].add(theme)

    ws = wb[SHEET]
    controls, title2idx = [], {}
    by_threat = collections.defaultdict(set)
    by_e = collections.defaultdict(set)

    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        # cols: 1 Threat cluster, 2 Threat, 3 Control title, 4 Control, 5 Control desc,
        #       6 Control cluster, 7 Reference frameworks, 8 Lifecycle, 9 Segment
        if len(r) < 8:
            continue
        cluster, threat, title = str(r[1] or "").strip(), str(r[2] or "").strip(), str(r[3] or "").strip()
        if not threat or not title or threat == "Threat":
            continue
        if title not in title2idx:
            frameworks = split(r[7])
            title2idx[title] = len(controls)
            controls.append({
                "title": title,
                "cluster": str(r[6] or "").strip(),
                "desc": str(r[5] or r[4] or "").strip()[:480],
                "frameworks": frameworks,
                "lifecycle": str(r[8] or "").strip(),
                "sparta": "SPARTA" in frameworks,
            })
        idx = title2idx[title]
        by_threat[threat.lower()].add(idx)
        e = CLUSTER_TO_E.get(cluster.lower())
        if e:
            by_e[e].add(idx)

    # associate each control with the SPARTA countermeasure themes its cluster covers
    for c in controls:
        th = set()
        for cp in cluster_parts(c["cluster"]):
            th |= cluster2themes.get(cp, set())
        c["themes"] = sorted(th)

    out = {
        "meta": {
            "source": "ENISA Space Threat Landscape — Control Framework",
            "repo": "https://github.com/enisaeu/Space-Threat-Landscape",
            "file": "Control Framework.xlsx",
        },
        "controls": controls,
        "byThreat": {k: sorted(v) for k, v in sorted(by_threat.items())},
        "byE": {k: sorted(v) for k, v in sorted(by_e.items())},
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(out, open(OUT, "w"), ensure_ascii=False, separators=(",", ":"))
    sparta_backed = sum(1 for c in controls if c["sparta"])
    with_themes = sum(1 for c in controls if c["themes"])
    print(f"controls={len(controls)} (SPARTA-referenced={sparta_backed}, "
          f"theme-linked={with_themes}) threats={len(by_threat)} "
          f"E-categories={len(by_e)} -> {os.path.relpath(OUT)}")
    print("E coverage:", {k: len(v) for k, v in sorted(by_e.items())})


if __name__ == "__main__":
    main()
