#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ntn_threat_algorithm.py  --  Algorithm 1: ENISA-SPARTA Integrated Threat Modeling for 6G NTN

Reads a threat-record tuple table (xlsx or csv) whose columns are the tuple
fields, runs the two-pass algorithm deterministically, and prints a report.

    tuple table  ->  python ntn_threat_algorithm.py <table>  ->  report

Pass 1 (ENISA, threat-centric):  the affected-asset SET A is the primitive;
                                 derive S (segments), TB (trust boundaries),
                                 L (lifecycle) from A and VALIDATE them.
Pass 2 (SPARTA):                 apply decision table R1-R10 to classify each
                                 threat, set the TTP arm (adversarial -> TTP
                                 expected; else -> N/A), and record CM/I/Conf
                                 status.  (SPARTA TTPs are mapped by the analyst
                                 and left blank here -> reported as PENDING.)
Post:                            flag R9 (multi-segment) / R10 (multi-lifecycle),
                                 coverage audit, and a summary.

Deterministic: no randomness, stable ordering -> identical output every run.
Usage:
    python ntn_threat_algorithm.py enisa_threats_tuple.xlsx
    python ntn_threat_algorithm.py enisa_threats_tuple.csv --out report.txt
"""
import sys, os, csv, argparse, re

# --------------------------------------------------------------------------- #
#  reference taxonomies (ENISA Space Threat Landscape 2025)
# --------------------------------------------------------------------------- #
LIFECYCLE = {1:"Design and Development",2:"Assembly",3:"Pre-launch",4:"Launch",
             5:"In-orbit Testing",6:"Operations",7:"Decommissioning"}
SEG_ORDER = ["User","Ground","Space","Human Resources"]

# category -> (nature, primary rule, TTP arm)
ADVERSARIAL = {"NAA","EIH","PA"}
FAILURE     = {"FM","UD"}
STATE       = {"OUT"}
ENVIRON     = {"DIS"}
ENABLING    = {"LEI","ACQ"}
LEGAL       = {"LEG"}

def classify_nature(code):
    if code in ADVERSARIAL: return ("adversarial",      "R1", "SPARTA TTP expected")
    if code in ENABLING:    return ("enabling/legacy",  "R5", "SPARTA TTP expected (spawns adversarial chain)")
    if code in STATE:       return ("state/outage",     "R3", "N/A (state kept as precondition)")
    if code in FAILURE:     return ("operational/unintentional","R2","N/A (non-adversarial)")
    if code in ENVIRON:     return ("physical/environmental","R4","N/A (non-adversarial)")
    if code in LEGAL:       return ("legal",            "--", "N/A (no SPARTA mapping)")
    return ("unknown","--","N/A")

# --------------------------------------------------------------------------- #
#  derivations (Algorithm 1, Pass 1)  --  mirror the workbook builder
# --------------------------------------------------------------------------- #
def derive_segments(affected):
    a = affected.lower(); segs=[]
    if "ground:" in a: segs.append("Ground")
    if "space:"  in a: segs.append("Space")
    if "user:"   in a: segs.append("User")
    if "human resources:" in a or "human resource:" in a: segs.append("Human Resources")
    # "Ground & Space:" style
    if "ground & space:" in a:
        for s in ("Ground","Space"):
            if s not in segs: segs.append(s)
    return [s for s in SEG_ORDER if s in segs]

def derive_trust_boundaries(segs):
    tbs=[]
    for a,b,label in [("User","Ground","User <-> Ground/RAN (service link)"),
                      ("Ground","Space","Ground <-> Space (feeder / telecommand link)"),
                      ("User","Space","User <-> Space (direct service link)")]:
        if a in segs and b in segs: tbs.append(label)
    if "Human Resources" in segs and len(segs)>1:
        tbs.append("Human Resources <-> technical segment (organizational boundary)")
    return tbs or ["(none - single segment / internal)"]

GROUND_EARLY={1,2,3}; GROUND_OPS={4,5,6,7}; SPACE_SPAN={3,4,5,6,7}; USER_SPAN={6}; HR_SPAN={1,2,3,4,5,6,7}
def derive_lifecycle(affected, segs):
    a=affected.lower(); ph=set()
    if "Ground" in segs:
        if "all assets" in a or any(k in a for k in ["ttc","control centre","mission control","crypto unit ground","operations"]):
            ph|=GROUND_OPS
        if "all assets" in a or any(k in a for k in ["production","design","development","assembly","manufactur","checkout","test","simulator","egse","mgse","document management","erp","prototyp","transport"]):
            ph|=GROUND_EARLY
        if "all assets" in a: ph|=GROUND_EARLY|GROUND_OPS
    if "Space" in segs: ph|=SPACE_SPAN
    if "User" in segs: ph|=USER_SPAN
    if "Human Resources" in segs: ph|=HR_SPAN
    return sorted(ph) or [6]

# --------------------------------------------------------------------------- #
#  input loading (xlsx or csv)
# --------------------------------------------------------------------------- #
FIELD_ALIASES = {
    "id":"ID","enisa_threat":"threat","threat":"threat",
    "e (category)":"E","e":"E","category":"E","cia":"CIA",
    "l (lifecycle)":"L","s (segments)":"S","a (affected assets)":"A",
    "tb (trust boundaries)":"TB","pre":"Pre","ttp":"TTP","cm":"CM","i":"I",
    "conf":"Conf","source (enisa stl 2025)":"Source",
}
def norm_headers(hdrs):
    return [FIELD_ALIASES.get((h or "").strip().lower(), (h or "").strip()) for h in hdrs]

def load_rows(path):
    ext=os.path.splitext(path)[1].lower()
    if ext in (".xlsx",".xlsm"):
        from openpyxl import load_workbook
        wb=load_workbook(path, data_only=True); ws=wb["Tuples"] if "Tuples" in wb.sheetnames else wb.active
        data=[[c.value for c in row] for row in ws.iter_rows()]
        hdr=norm_headers(data[0]); body=data[1:]
    else:
        with open(path, encoding="utf-8-sig") as f:
            r=list(csv.reader(f)); hdr=norm_headers(r[0]); body=r[1:]
    rows=[]
    for raw in body:
        if not any(raw): continue
        rows.append({hdr[i]: ("" if i>=len(raw) or raw[i] is None else str(raw[i]).strip())
                     for i in range(len(hdr))})
    return rows

# --------------------------------------------------------------------------- #
#  Algorithm 1
# --------------------------------------------------------------------------- #
def parse_phases(Lstr):
    return sorted(int(n) for n in re.findall(r'(\d+):', Lstr or ""))

def run(rows):
    out=[]
    for row in rows:
        code=(row.get("E","").split("/")[0]).strip()
        A=row.get("A",""); tid=row.get("ID","?"); name=row.get("threat","?")
        # ---- Pass 1: derive from A ----
        dS=derive_segments(A); dTB=derive_trust_boundaries(dS); dL=derive_lifecycle(A,dS)
        # validate vs stored
        sS=[s.strip() for s in re.split(r'[;,]', row.get("S","")) if s.strip()]
        s_ok = ([s for s in SEG_ORDER if s in sS]==dS) if sS else None
        # ---- Pass 2: decision table ----
        nature,prim,arm=classify_nature(code)
        rules=set()
        if code in ADVERSARIAL: rules.add("R1")
        if code in FAILURE:     rules.add("R2")
        if code in STATE:       rules.add("R3")
        if code in ENVIRON:     rules.add("R4")
        if code in ENABLING:    rules.add("R5")
        adversarial = code in ADVERSARIAL or code in ENABLING
        if adversarial: rules|={"R6","R8"}       # 1<->many techniques; CM linkage
        rules.add("R7")                            # Conf tagging always
        if len(dS)>=2: rules.add("R9")            # multi-segment footprint
        if len(dL)>=2: rules.add("R10")           # multi-lifecycle
        # ---- TTP/CM/Conf status (SPARTA side blank) ----
        ttp=row.get("TTP","").strip()
        if arm.startswith("N/A"):
            ttp_status="N/A (by rule "+prim+")"
        else:
            ttp_status=("MAPPED: "+ttp) if ttp else "PENDING (analyst to map SPARTA TTPs)"
        conf=row.get("Conf","").strip() or ("--" if arm.startswith("N/A") else "PENDING")
        out.append(dict(id=tid,name=name,code=code,cia=row.get("CIA",""),
                        nature=nature,arm=arm,prim=prim,rules=sorted(rules,key=_rk),
                        dS=dS,dTB=dTB,dL=dL,s_ok=s_ok,A=A,
                        ttp_status=ttp_status,conf=conf,
                        multiseg=len(dS)>=2,multilife=len(dL)>=2))
    return out

def _rk(r):  # rule sort key R1..R10
    return int(r[1:]) if r[1:].isdigit() else 99

# --------------------------------------------------------------------------- #
#  report
# --------------------------------------------------------------------------- #
def report(results, src):
    L=[]; W=L.append
    bar="="*78
    W(bar)
    W("  6G NTN THREAT MODELING  --  Algorithm 1 (ENISA x SPARTA) RESULT REPORT")
    W("  tuple: TR = <L, S, A, TB, E, Pre, TTP, CM, I, Conf>")
    W("  input: "+os.path.basename(src))
    W(bar)
    W("")
    n=len(results)
    adv=[r for r in results if r["arm"]=="SPARTA TTP expected" or "expected" in r["arm"]]
    na =[r for r in results if r["arm"].startswith("N/A")]
    ms =[r for r in results if r["multiseg"]]
    ml =[r for r in results if r["multilife"]]
    incons=[r for r in results if r["s_ok"] is False]
    W("SUMMARY")
    W("-"*78)
    W(f"  threat records processed .......... {n}")
    W(f"  adversarial (TTP arm, R1/R5) ...... {len(adv):>3}   -> SPARTA mapping applies")
    W(f"  non-adversarial (TTP = N/A) ....... {len(na):>3}   -> state/failure/env/legal (R2/R3/R4)")
    W(f"  multi-segment footprint (R9) ...... {len(ms):>3}   -> bridging sub-chain within one record")
    W(f"  multi-lifecycle (R10) ............. {len(ml):>3}")
    W(f"  Pass-1 S-derivation consistent .... {n-len(incons)}/{n}"
      + ("" if not incons else f"   (!! {len(incons)} MISMATCH)"))
    # category breakdown
    cats={}
    for r in results: cats[r["code"]]=cats.get(r["code"],0)+1
    W("  by ENISA category (E) ............. "+", ".join(f"{k}:{v}" for k,v in sorted(cats.items())))
    W("")

    # per-threat detail
    W("PER-THREAT RESULTS")
    W("="*78)
    for r in results:
        W(f"[{r['id']}]  {r['name']}")
        W(f"     E (category)  : {r['code']}   CIA: {r['cia']}   nature: {r['nature']}")
        W(f"     Pass 1  S     : {{{', '.join(r['dS'])}}}"
          + ("   [OK derived from A]" if r["s_ok"] else ("   [!! differs from stored S]" if r["s_ok"] is False else "")))
        W(f"             TB    : " + "; ".join(r["dTB"]))
        W(f"             L     : {{{', '.join(str(p)+':'+LIFECYCLE[p] for p in r['dL'])}}}")
        W(f"     Pass 2  rules : {', '.join(r['rules'])}   (primary {r['prim']})")
        W(f"             TTP   : {r['ttp_status']}")
        flags=[]
        if r["multiseg"]: flags.append("R9 multi-segment -> emit bridging sub-chain IA->EX/LM->IMP across "+"/".join(r["dS"]))
        if r["multilife"]: flags.append(f"R10 multi-lifecycle -> L spans {len(r['dL'])} phases (split only if TTP/CM differ)")
        if r["code"]=="OUT": flags.append("R3 state -> record as enabling precondition; adversarial pair is a separate record")
        for fl in flags: W(f"             flag  : {fl}")
        W(f"             CM/I/Conf : CM=PENDING  I=PENDING  Conf={r['conf']}")
        W("")

    # chaining / audit note
    W("POST-PASS  (Algorithm 1, lines 20-21)")
    W("-"*78)
    W("  * Attack-path chaining (line 20): links records where a post-condition")
    W("    satisfies another record's precondition. Requires the analyst's Pre/TTP")
    W("    fields -> run again after SPARTA mapping to materialize multi-stage paths.")
    W("  * Coverage audit (line 21): every (asset, phase) with an ENISA lifecycle")
    W("    entry should be covered by >=1 record; gaps feed the evaluation (sec. 8.2).")
    W("")
    W("NEXT STEP")
    W("-"*78)
    W("  Fill Pre / TTP / CM / I / Conf for the "+str(len(adv))+" adversarial records"
      " (SPARTA v3.2),")
    W("  then re-run: the report will show MAPPED TTPs, CM crosswalks, and chained paths.")
    W(bar)
    return "\n".join(L)

# --------------------------------------------------------------------------- #
def main():
    ap=argparse.ArgumentParser(description="Algorithm 1: ENISA-SPARTA 6G NTN threat modeling")
    ap.add_argument("table", help="tuple table (.xlsx or .csv)")
    ap.add_argument("--out", help="also write the report to this file")
    a=ap.parse_args()
    if not os.path.exists(a.table):
        sys.exit(f"error: file not found: {a.table}")
    rows=load_rows(a.table)
    results=run(rows)
    text=report(results, a.table)
    print(text)
    if a.out:
        open(a.out,"w",encoding="utf-8").write(text+"\n")
        print(f"\n[written] {a.out}")

if __name__=="__main__":
    main()
